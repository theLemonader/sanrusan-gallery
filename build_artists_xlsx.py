#!/usr/bin/env python3
"""Build SANRUSAN artists outreach spreadsheet with embedded thumbnails."""

import os
from pathlib import Path
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

BASE = Path("/Users/san/sanrusan/artisti")
OUT_THUMBS = BASE / "tiles"  # use the cropped work-tiles
OUT_THUMBS.mkdir(exist_ok=True)
OUT_XLSX = Path("/Users/san/sanrusan/SANRUSAN_artists.xlsx")

# ---------- ARTIST DATABASE (consolidated from screenshots + web research) ----------
ARTISTS = [
    {
        "img": "IMG_9633.PNG",
        "name": "Meike & Johannes Budde",
        "studio": "BUDDE",
        "handle": "@johannesbudde",
        "website": "budde.co",
        "email": "meike@budde.co",
        "phone": "+49 176 41326982",
        "location": "Cologne, Germany (also Milan)",
        "address": "Bachemer Straße 121, 50931 Köln",
        "tags": "Furniture, lighting, interior objects; collectible + commercial",
        "approach": "Email Meike (business lead). Also met by martamarghidanu, bogdantutuneanu (mutual followers).",
    },
    {
        "img": "IMG_9634.PNG",
        "name": "DIVE Design",
        "studio": "DIVE",
        "handle": "@divedesignco",
        "website": "divedesignco.com / additive.divedesignco.com",
        "email": "via website",
        "phone": "",
        "location": "Boonton, NJ, USA",
        "address": "",
        "tags": "Additive manufacturing / 3D printing; sustainable furniture; DECIBEL editions; 3D Pets",
        "approach": "Pitch via additive.divedesignco.com — Alex (CEO) & Adam Hecht (Director AM). Mention DECIBEL editions / collectible angle.",
    },
    {
        "img": "IMG_9635.PNG",
        "name": "Giulio Camposano",
        "studio": "Giulio Pottery Studio",
        "handle": "@giuliopotterystudio",
        "website": "giuliopotterystudio.art",
        "email": "Instagram DM",
        "phone": "",
        "location": "Verbania, Italy",
        "address": "",
        "tags": "Brutalist ceramics; Japanese minimalism; pottery wheel, stoneware",
        "approach": "DM via Instagram (most active). Mutual: the_1of1_. Mention SANRUSAN gallery curation.",
    },
    {
        "img": "IMG_9636.PNG",
        "name": "form & eisen",
        "studio": "form & eisen",
        "handle": "@form.eisen",
        "website": "form-eisen.at",
        "email": "office@form-eisen.at",
        "phone": "+43 664 2387708",
        "location": "Salzburg, Austria",
        "address": "",
        "tags": "Collectible steel furniture; sideboards, shelves, stools, planters; industrial elegance",
        "approach": "Email office@. Mutual: biju.gallery. Worldwide delivery already; should be receptive.",
    },
    {
        "img": "IMG_9637.PNG",
        "name": "Balboa Design",
        "studio": "Balboa Design",
        "handle": "@balboa.design",
        "website": "balboadesign.fr",
        "email": "via website contact form",
        "phone": "",
        "location": "France",
        "address": "",
        "tags": "Design studio - furniture & objects, made in-house; PSILO / BESPOKE / WORKSHOP / EVENTS",
        "approach": "Contact via balboadesign.fr. Strong sculptural furniture in metal — fits SANRUSAN aesthetic.",
    },
    {
        "img": "IMG_9638.PNG",
        "name": "Casa Simples DIY",
        "studio": "Casa Simples",
        "handle": "@casasimplesdiy",
        "website": "",
        "email": "WhatsApp + IG DM",
        "phone": "",
        "location": "Brazil (Campinas / São Paulo area implied)",
        "address": "",
        "tags": "DIY, decor inspiration, achadinhos (small finds), functional decor",
        "approach": "WhatsApp link in bio + IG DM. More of a creator/influencer than studio — could be partnership for visibility.",
    },
    {
        "img": "IMG_9639.PNG",
        "name": "Luca Ricci",
        "studio": "Luca Ricci",
        "handle": "@lucaa.ricci",
        "website": "behance.net/lucaaricci",
        "email": "IG DM / via Behance",
        "phone": "",
        "location": "Milan, Italy",
        "address": "",
        "tags": "Art Director / Graphic Designer; experimental objects; collectible art-objects",
        "approach": "DM Instagram; portfolio on Behance. Mutual: marcosimonetti____.",
    },
    {
        "img": "IMG_9640.PNG",
        "name": "Benjamin Bacon (白培耕)",
        "studio": "SoundSpade / Dogma Lab",
        "handle": "@soundspade",
        "website": "benjaminbacon.studio",
        "email": "via website contact",
        "phone": "",
        "location": "Shanghai, China",
        "address": "",
        "tags": "Artist / Associate Professor Media & Arts @ Duke Kunshan; sound, kinetic, experimental",
        "approach": "Contact via benjaminbacon.studio. Pitch as functional art / sound objects intersection.",
    },
    {
        "img": "IMG_9641.PNG",
        "name": "Felipe Pantone",
        "studio": "Felipe Pantone Studio (FPSTUDIO)",
        "handle": "@felipepantone",
        "website": "felipepantone.com",
        "email": "via felipepantone.com/contact",
        "phone": "",
        "location": "Valencia, Spain",
        "address": "",
        "tags": "Kinetic / chromatic / configurable art; collectibles via @configurableart, @casaaxis",
        "approach": "Huge artist (771K). Better path: via @configurableart shop or DM @casaaxis (his gallery brand). 771K following — long shot but viral fit.",
    },
    {
        "img": "IMG_9642.PNG",
        "name": "Laurids Gallée",
        "studio": "Studio Laurids Gallée",
        "handle": "@laurids.gallee",
        "website": "lauridsgallee.com",
        "email": "via website (unique pieces inquiries)",
        "phone": "",
        "location": "Rotterdam, NL (Viennese)",
        "address": "",
        "tags": "Resin furniture, unique pieces or small series; rep'd by Objects With Narratives, The Future Perfect",
        "approach": "Email via lauridsgallee.com. Mutual followers: martamarghidanu, ozymandias.inc. Strong SANRUSAN fit.",
    },
    {
        "img": "IMG_9643.PNG",
        "name": "Daniel de Bruin",
        "studio": "Daniel de Bruin",
        "handle": "@daniel_de_bruin",
        "website": "danieldebruin.com",
        "email": "via website",
        "phone": "",
        "location": "Utrecht / Soesterberg, NL",
        "address": "",
        "tags": "Creative engineer; prototypical machines, devices, kinetic objects",
        "approach": "Contact via danieldebruin.com. Followed by felipepantone, fabianoefner. Beloved of design press.",
    },
    {
        "img": "IMG_9644.PNG",
        "name": "Onibi",
        "studio": "Onibi",
        "handle": "@oni.bi",
        "website": "onibi.us",
        "email": "via website / IG DM",
        "phone": "",
        "location": "Brooklyn, NY, USA",
        "address": "",
        "tags": "Kinetic instruments for illumination; Ramulus floor lamp; collectible lighting",
        "approach": "Email through onibi.us product/contact page. Mutual: demauricio_. Perfect functional-art fit.",
    },
    {
        "img": "IMG_9645.PNG",
        "name": "KALA arts + objects",
        "studio": "KALA",
        "handle": "@kala_artobjects",
        "website": "kalaartobjects.com",
        "email": "support@kalaartobjects.com",
        "phone": "+91 9946634891",
        "location": "India",
        "address": "",
        "tags": "Room accents, soft furnishings, wall art, rugs, lighting; ships worldwide",
        "approach": "Email support@. Possibly e-commerce reseller more than original artist — useful for stock/dropship discussions.",
    },
    {
        "img": "IMG_9646.PNG",
        "name": "DAKD Jung",
        "studio": "BurnSlap",
        "handle": "@dakd_jung",
        "website": "burnslap.me",
        "email": "via burnslap.me subscription / IG DM",
        "phone": "",
        "location": "Seoul, South Korea",
        "address": "",
        "tags": "Tech-based installation art; collectible electronics / FAV-EQ; release-by-application drops",
        "approach": "Subscribe on burnslap.me to receive release news; DM @dakd_jung. Mutual: ozymandias.inc, iamrurd.",
    },
    {
        "img": "IMG_9647.PNG",
        "name": "Andrea Mancuso",
        "studio": "Studio Andrea Mancuso",
        "handle": "@andreamancuso_studio",
        "website": "andreamancuso.com",
        "email": "studio@andreamancuso.com",
        "phone": "",
        "location": "Milan, Italy",
        "address": "Via Marco Aurelio 21, 20127 Milano",
        "tags": "Research-based; narrative-driven objects, furniture, installations; AD100; Elle Deco winner 2025",
        "approach": "Email studio@. Mutual: valentinrusan, objectswithnarratives. Top-tier — pitch as exclusive Eastern European debut.",
    },
    {
        "img": "IMG_9648.PNG",
        "name": "Draga Obradovic & Aurel K. Basedow",
        "studio": "Draga & Aurel",
        "handle": "@dragaandaurel",
        "website": "draga-aurel.com",
        "email": "via website contact form",
        "phone": "+39 031 3370189 (gallery) / +39 031 522817 (studio)",
        "location": "Como, Italy",
        "address": "Atelier: Via G. Ferrari 7, 22100 Como; Studio: Via A. Lenticchia 15, 22100 Como",
        "tags": "Art & design atelier; resin, mixed media; #dragaandaurel #transparencymatters",
        "approach": "Call or email. Mutual: martamarghidanu, valentinrusan. Established (45K followers).",
    },
    {
        "img": "IMG_9649.PNG",
        "name": "Moritz Waldemeyer",
        "studio": "Studio Waldemeyer",
        "handle": "@studio.waldemeyer",
        "website": "studio.waldemeyer.com",
        "email": "studio@waldemeyer.com",
        "phone": "",
        "location": "London, UK",
        "address": "",
        "tags": "Art Engineering & Light Design; clients: Philip Treacy, Bentley, Jamiroquai",
        "approach": "Email studio@. Mutual: valentinrusan. Celebrity-tier creative engineer.",
    },
    {
        "img": "IMG_9650.PNG",
        "name": "Deniz Aktay",
        "studio": "Dezin Objects",
        "handle": "@dezinobjects",
        "website": "dezinobjects.com",
        "email": "info@dezinobjects.com",
        "phone": "",
        "location": "Stuttgart, Germany",
        "address": "",
        "tags": "Furniture & object design; press & awards listed; functionalist sculptural pieces",
        "approach": "Email info@. Mutual: buzzyalex, andrada.borda. Active commercial — likely open to gallery placement.",
    },
    {
        "img": "IMG_9651.PNG",
        "name": "Vincenzo De Cotiis",
        "studio": "Vincenzo De Cotiis Architects & Gallery",
        "handle": "@vdecotiis",
        "website": "decotiis.it",
        "email": "gallery@decotiis.it",
        "phone": "",
        "location": "Milan, Italy",
        "address": "Via San Giovanni sul muro 18, 20121 Milano",
        "tags": "Art / Architecture / Interior Design; ART BASEL, Venice; ultra high-end collectible",
        "approach": "Email gallery@. Mutual: teodoraburz, martamarghidanu. Top-tier — long shot but worth a curated outreach.",
    },
    {
        "img": "IMG_9652.PNG",
        "name": "Arielle Assouline-Lichten",
        "studio": "Slash Objects",
        "handle": "@slashobjects",
        "website": "slashobjects.com",
        "email": "via slashobjects.com / typeform link in bio",
        "phone": "",
        "location": "New York + Paris",
        "address": "224 Centre Street, 4th Floor, New York 10013",
        "tags": "Furniture, surfaces, rubber/cork experimentation; Milan 2026; X Series",
        "approach": "Fill typeform from her bio; mutual: ozymandias.inc, valentinrusan. Material-first practice — strong SANRUSAN match.",
    },
    {
        "img": "IMG_9653.PNG",
        "name": "TARZ",
        "studio": "TARZ Hi-Fi Systems",
        "handle": "@tarzhifi",
        "website": "",
        "email": "Instagram DM (no website found)",
        "phone": "",
        "location": "Brazil (likely)",
        "address": "",
        "tags": "High-fidelity sound systems, custom-built, design-driven; BOLT B6 / BOLT Series",
        "approach": "Only IG (88 followers — very new). DM directly. Could be founding partnership opportunity.",
    },
    {
        "img": "IMG_9654.PNG",
        "name": "Javier & Anaïs",
        "studio": "UNAVELA",
        "handle": "@unavela.store",
        "website": "unavela.store",
        "email": "via unavela.store",
        "phone": "",
        "location": "South of France",
        "address": "",
        "tags": "Everyday objects with unusual shapes; aluminium machined; aerospace-engineer founders",
        "approach": "Contact via unavela.store. Mutual: diamantby, ovidiubojor. Recent Yanko press — momentum.",
    },
    {
        "img": "IMG_9655.PNG",
        "name": "AOT studio",
        "studio": "AOT",
        "handle": "@a.o.t.studio",
        "website": "aotstudio.com",
        "email": "via aotstudio.com contact",
        "phone": "",
        "location": "Copenhagen, Denmark",
        "address": "",
        "tags": "Family-led Danish design; one-off pieces + collections; 3daysofdesign exhibitor",
        "approach": "Contact via aotstudio.com. Mutual: laurids.gallee, anneivan. Established Danish — quality fit.",
    },
    {
        "img": "IMG_9656.PNG",
        "name": "Nicholas Baker",
        "studio": "Nicholas Baker Studio",
        "handle": "@nickpbaker",
        "website": "baker.studio",
        "email": "via baker.studio / Substack",
        "phone": "",
        "location": "Brooklyn, NY, USA",
        "address": "",
        "tags": "Furniture, lighting, everyday objects; The Local Project profile",
        "approach": "Contact via baker.studio. Mutual: raressmanolache, cozmajohn. Subscribe to bakerstudio.substack.com for inroad.",
    },
    {
        "img": "IMG_9657.PNG",
        "name": "Alberto Esses",
        "studio": "Essesi Design Studio",
        "handle": "@albertoessesi",
        "website": "essesi.com / essesi.design",
        "email": "hello@essesi.com",
        "phone": "",
        "location": "Los Angeles, USA (Mexican-born)",
        "address": "",
        "tags": "Industrial design; former Tesla Design Studio lead; Mexican craftsmanship + tech",
        "approach": "Email hello@. Mutual: andrada.borda, raressmanolache. 'Sentience to Matter' philosophy aligns with SANRUSAN.",
    },
    {
        "img": "IMG_9658.PNG",
        "name": "Paul Cocksedge",
        "studio": "Paul Cocksedge Studio",
        "handle": "@paulcocksedge",
        "website": "paulcocksedgestudio.com",
        "email": "info@paulcocksedgestudio.com",
        "phone": "+44 20 8985 0907",
        "location": "London, UK",
        "address": "2A Brenthouse Road, Soloman's Yard, London E9 6QG",
        "tags": "Major artist & designer; collaborations with Dior; ships worldwide; PCS/NEWS/DROPS",
        "approach": "Email info@. Mutual: martamarghidanu, bogzafmir. 613K followers — huge name but he sells limited drops.",
    },
    {
        "img": "IMG_9659.PNG",
        "name": "Giacomo Ravagli",
        "studio": "Giacomo Ravagli",
        "handle": "@giacomo_ravagli",
        "website": "giacomoravagli.com",
        "email": "via giacomoravagli.com",
        "phone": "",
        "location": "Pietrasanta, Italy",
        "address": "",
        "tags": "Marble carver; high-end furniture, unique pieces, limited editions; Carpenters Workshop Gallery",
        "approach": "Email via website. Mutual: valentinrusan, nilufargallery. Gallery-represented — go via him directly first.",
    },
    {
        "img": "IMG_9660.PNG",
        "name": "Lukas Cober",
        "studio": "Studio Lukas Cober",
        "handle": "@studiolukascober",
        "website": "lukascober.com",
        "email": "via lukascober.com",
        "phone": "",
        "location": "Maastricht, Netherlands",
        "address": "Sandersweg 33, 6219 NW Maastricht",
        "tags": "Makers Studio — Shapes — Unique Pieces; sculptural resin/glass furniture",
        "approach": "Contact via lukascober.com. Mutual: laurids.gallee, diamantby. Strong SANRUSAN fit — confirmed gallery contacts.",
    },
    {
        "img": "IMG_9661.PNG",
        "name": "Manu Bañó (Manuel)",
        "studio": "Manu Bañó",
        "handle": "@manuelbano",
        "website": "manubano.com",
        "email": "via manubano.com",
        "phone": "",
        "location": "Valencia, Spain / CDMX, Mexico",
        "address": "",
        "tags": "Co-founder EWE Studio; rep'd by @gallery_collectional, @licht_gallery, @masagaleria, @objectswithnarratives",
        "approach": "Contact via manubano.com. Mutual: martamarghidanu, laurids.gallee. Heavy gallery rep — pitch quality of curation.",
    },
    {
        "img": "IMG_9662.PNG",
        "name": "Héctor Esrawe",
        "studio": "Esrawe Studio",
        "handle": "@hesrawe",
        "website": "esrawe.com",
        "email": "studio@esrawe.com (press: medios@esrawe.com)",
        "phone": "+52 55 5553 9611",
        "location": "Mexico City, Mexico",
        "address": "Córdoba 206, Roma Norte, CDMX",
        "tags": "HFAIA Director; co-founder @ewestudio, @xinuperfumes, @masagaleria; large architecture+design practice",
        "approach": "Email studio@ or medios@. Mutual: fabianoefner, diamantby. Major studio — likely via @masagaleria pipeline.",
    },
    {
        "img": "IMG_9663.PNG",
        "name": "Vladimir Slavov",
        "studio": "DIM atelier",
        "handle": "@dimatelier",
        "website": "dimatelier.com",
        "email": "vladimir.slavov@dimatelier.com",
        "phone": "+32 484 327 077",
        "location": "Antwerp / Zaventem, Belgium",
        "address": "",
        "tags": "Custom lighting and objects; PAD London, Collectible Fair, Nuance",
        "approach": "Email Vladimir directly. Mutual: laurids.gallee, diamantby. Listed in 'Objects With Narratives'. Hot fit.",
    },
    {
        "img": "IMG_9664.PNG",
        "name": "hasik design studio",
        "studio": "hasik design studio",
        "handle": "@hasik_design",
        "website": "hasik.design",
        "email": "contact@hasik.design",
        "phone": "+48 881 321 311 / +48 514 833 987",
        "location": "Poland",
        "address": "",
        "tags": "Unique, handcrafted objects + authentic interior design; COMURE series",
        "approach": "Email contact@. Polish — geographic proximity to Romania useful for shipping/visits.",
    },
    {
        "img": "IMG_9665.PNG",
        "name": "Arthur Moulucou",
        "studio": "MLK Furniture",
        "handle": "@mlkfurniture",
        "website": "MLK on 1stDibs, Etsy, APOC, Isola, basic.space",
        "email": "arthur.moulucou@mailfence.com",
        "phone": "+33 6 12 98 01 81",
        "location": "Saint-Ciers-sur-Gironde, France",
        "address": "114 Avenue de la République, 33820 Saint-Ciers-sur-Gironde",
        "tags": "Functionalism + brutalism; low-tech, ready-to-assemble, modular; Hex400 table",
        "approach": "Email arthur directly. Mutual: diamantby, super.sedia. Strong philosophical alignment with SANRUSAN.",
    },
    {
        "img": "IMG_9666.PNG",
        "name": "Guillaume Crédoz",
        "studio": "Bits to Atoms / Post Industrial Crafts",
        "handle": "@ghouyoum",
        "website": "bitstoatoms.xyz",
        "email": "via bitstoatoms.xyz contact",
        "phone": "",
        "location": "M'kalles, Beirut, Lebanon",
        "address": "",
        "tags": "Architect & designer; bike trips, processes, sketches; Architecture Masterprize winner",
        "approach": "Contact via bitstoatoms.xyz. Lebanese architecture office — uncommon geography, distinctive output.",
    },
    {
        "img": "IMG_9667.PNG",
        "name": "Shivas Howard-Brown",
        "studio": "Friendly Pressure",
        "handle": "@friendly.pressure",
        "website": "friendlypressure.studio",
        "email": "Email via Instagram bio (for all enquiries)",
        "phone": "",
        "location": "London, UK",
        "address": "",
        "tags": "Bespoke loudspeaker systems; horn-inspired; Stone Island collab; FP-498, FP-Y210",
        "approach": "Email via bio link. Mutual: valentinrusan, anaraucea. Crossover fashion/audio — high-design audience.",
    },
    {
        "img": "IMG_9668.PNG",
        "name": "Drew Hart",
        "studio": "East 38th",
        "handle": "@east38th_",
        "website": "east38th.com",
        "email": "via east38th.com",
        "phone": "",
        "location": "Cincinnati, OH, USA",
        "address": "",
        "tags": "Furniture designer & maker; sculptural wood; Ripple Wall lamp",
        "approach": "Contact via east38th.com. Mutual: demauricio_. Heavy maker presence (2M views on a Reel) — bankable craft.",
    },
    {
        "img": "IMG_9669.PNG",
        "name": "Romanovna",
        "studio": "Romanovna Studio",
        "handle": "@romanovna.std",
        "website": "admiddleeast.com feature / IG bio",
        "email": "IG DM",
        "phone": "",
        "location": "Middle East (per AD Middle East feature)",
        "address": "",
        "tags": "Functional Art Pieces, Limited Works, Sensory & Contemporary; AI-driven; 'coming to life very soon'",
        "approach": "DM Instagram. Mutual: tudormonroe, avi_ciciream. Pre-launch — perfect timing for SANRUSAN co-launch.",
    },
    {
        "img": "IMG_9670.PNG",
        "name": "Brodie Neill",
        "studio": "Brodie Neill (Made in Ratio)",
        "handle": "@brodieneill",
        "website": "brodieneill.com",
        "email": "info@brodieneill.com",
        "phone": "+44 20 7033 3434",
        "location": "London, UK",
        "address": "4B Orsman Road, Tuscany Wharf, London N1 5QJ",
        "tags": "Collectable & limited-edition; material innovation, sculptural form; SHOWING: Opal at One Leadenhall",
        "approach": "Email info@. Mutual: objectswithnarratives, demauricio_. Established collectible — credibility booster for SANRUSAN.",
    },
]

# ---------- ATTACH TILE PATHS (already cropped by crop_tiles.py) ----------
print(f"Attaching {len(ARTISTS)} tile paths…")
for a in ARTISTS:
    dst = OUT_THUMBS / a["img"].replace(".PNG", ".jpg")
    a["thumb"] = str(dst)
    if not dst.exists():
        print(f"  WARN: missing tile {dst}; run crop_tiles.py first")

# ---------- BUILD XLSX ----------
print("Building xlsx…")
wb = Workbook()
ws = wb.active
ws.title = "SANRUSAN Outreach"

# Style tokens
HEADER_FONT = Font(name="Helvetica Neue", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="111111")
BODY_FONT = Font(name="Helvetica Neue", size=10)
LINK_FONT = Font(name="Helvetica Neue", size=10, color="0066CC", underline="single")
CENTER = Alignment(horizontal="left", vertical="top", wrap_text=True)
THIN = Side(border_style="thin", color="DDDDDD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLS = [
    ("Image", 26),
    ("Artist / Maker", 24),
    ("Studio", 22),
    ("Instagram", 22),
    ("Website", 28),
    ("Email", 32),
    ("Phone", 22),
    ("Location", 24),
    ("Address", 32),
    ("Tags / Specialty", 38),
    ("Approach Notes", 46),
]

# Header
for col_idx, (label, width) in enumerate(COLS, start=1):
    cell = ws.cell(row=1, column=col_idx, value=label)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = BORDER
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.row_dimensions[1].height = 26
ws.freeze_panes = "B2"

# Data rows — tiles are 430x430 square; embed at 180x180
ROW_HEIGHT = 145  # accommodates 180px image
for r_idx, a in enumerate(ARTISTS, start=2):
    ws.row_dimensions[r_idx].height = ROW_HEIGHT

    # Image
    try:
        img = XLImage(a["thumb"])
        img.width = 180
        img.height = 180
        cell_anchor = f"A{r_idx}"
        ws.add_image(img, cell_anchor)
    except Exception as e:
        ws.cell(row=r_idx, column=1, value=f"(thumb error: {e})")

    values = [
        "",  # image placeholder column
        a.get("name", ""),
        a.get("studio", ""),
        a.get("handle", ""),
        a.get("website", ""),
        a.get("email", ""),
        a.get("phone", ""),
        a.get("location", ""),
        a.get("address", ""),
        a.get("tags", ""),
        a.get("approach", ""),
    ]
    for c_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=r_idx, column=c_idx, value=val)
        cell.font = BODY_FONT
        cell.alignment = CENTER
        cell.border = BORDER

        # Make website/instagram links clickable
        if c_idx == 4 and val and val.startswith("@"):
            cell.hyperlink = f"https://instagram.com/{val.lstrip('@')}"
            cell.font = LINK_FONT
        elif c_idx == 5 and val and "." in val and " " not in val:
            url = val if val.startswith("http") else f"https://{val.split('/')[0]}"
            cell.hyperlink = url
            cell.font = LINK_FONT
        elif c_idx == 6 and val and "@" in val and " " not in val and "via" not in val.lower() and "instagram" not in val.lower():
            cell.hyperlink = f"mailto:{val}"
            cell.font = LINK_FONT

# Subtle row striping for legibility
STRIPE = PatternFill("solid", fgColor="FAFAFA")
for r_idx in range(2, len(ARTISTS) + 2):
    if r_idx % 2 == 0:
        for c_idx in range(1, len(COLS) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            if not cell.fill or cell.fill.fgColor.rgb in (None, "00000000"):
                cell.fill = STRIPE

wb.save(OUT_XLSX)
print(f"\nDONE: {OUT_XLSX}")
print(f"Artists: {len(ARTISTS)}")
print(f"Thumbnails: {OUT_THUMBS}")
